"""Create an audit-friendly Excel workbook from the SMEE SQLite database."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from app.repositories import Database

HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
HIGH_FILL = PatternFill("solid", fgColor="FCE4D6")
MEDIUM_FILL = PatternFill("solid", fgColor="FFF2CC")
REVIEW_FILL = PatternFill("solid", fgColor="E4DFEC")
TABLE_STYLE = "TableStyleMedium2"


class ExcelReportExporter:
    """Export events and their audit trail to a formatted XLSX workbook."""

    sheet_names = ("Eventos", "Publicaciones", "Actores", "Revisión manual", "Fuentes")

    def __init__(self, database: Database) -> None:
        self.database = database

    def export(self, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name in self.sheet_names:
            workbook.create_sheet(name)

        with self.database.connect() as connection:
            publication_rows = connection.execute(
                """SELECT p.id, e.event_code, s.name AS source_name, p.title, p.published_at,
                          p.state_detected, p.municipality_detected, p.party_detected,
                          p.event_type_detected, p.status, p.needs_review,
                          ep.relationship_type, ep.is_primary_source,
                          CASE WHEN p.status='duplicate' THEN 1 ELSE 0 END AS is_duplicate,
                          p.url, p.author
                   FROM publications p
                   JOIN sources s ON s.id=p.source_id
                   LEFT JOIN event_publications ep ON ep.publication_id=p.id
                   LEFT JOIN events e ON e.id=ep.event_id
                   ORDER BY p.published_at DESC, p.id DESC"""
            ).fetchall()
            event_rows = connection.execute(
                """SELECT e.id, e.event_code, e.title, e.description, e.event_type, e.state,
                          e.municipality, e.start_date, e.last_update, e.status,
                          e.importance_level, e.priority_score, e.needs_review,
                          GROUP_CONCAT(DISTINCT a.name) AS actors, e.score_reasons
                   FROM events e
                   LEFT JOIN event_actors ea ON ea.event_id=e.id
                   LEFT JOIN actors a ON a.id=ea.actor_id
                   WHERE e.status != 'discarded'
                   GROUP BY e.id ORDER BY e.priority_score DESC, e.last_update DESC"""
            ).fetchall()
            actor_rows = connection.execute(
                """SELECT e.event_code, e.title AS event_title, a.name, a.actor_type,
                          a.party, a.state, a.is_priority
                   FROM event_actors ea JOIN events e ON e.id=ea.event_id
                   JOIN actors a ON a.id=ea.actor_id
                   ORDER BY e.event_code, a.name"""
            ).fetchall()
            review_rows = connection.execute(
                """SELECT p.id, p.title, s.name AS source_name, p.state_detected,
                          p.event_type_detected, p.review_reasons, p.url
                   FROM publications p JOIN sources s ON s.id=p.source_id
                   WHERE p.needs_review=1 ORDER BY p.collected_at DESC"""
            ).fetchall()
            source_rows = connection.execute(
                """SELECT id, name, source_type, state, reliability_level, is_active, base_url
                   FROM sources ORDER BY name"""
            ).fetchall()

        publication_end = max(2, len(publication_rows) + 1)
        self._write_publications(workbook["Publicaciones"], publication_rows)
        self._write_events(workbook["Eventos"], event_rows, publication_end)
        self._write_actors(workbook["Actores"], actor_rows)
        self._write_review(workbook["Revisión manual"], review_rows)
        self._write_sources(workbook["Fuentes"], source_rows, publication_end)

        workbook.properties.title = "SMEE - Monitoreo electoral"
        workbook.properties.subject = "Eventos, publicaciones y trazabilidad electoral"
        workbook.properties.creator = "SMEE"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(output_path)
        self.validate(output_path)
        return output_path

    def validate(self, path: Path) -> None:
        """Reopen the workbook and verify sheets, formulas, URLs, and dimensions."""
        workbook = load_workbook(path, data_only=False, read_only=False)
        if tuple(workbook.sheetnames) != self.sheet_names:
            raise ValueError("El archivo Excel no contiene las hojas esperadas")
        for sheet in workbook.worksheets:
            if sheet.max_row < 2 or sheet.max_column < 1:
                raise ValueError(f"La hoja {sheet.title} no contiene una tabla legible")
            if sheet.freeze_panes != "A2" or not sheet.auto_filter.ref:
                raise ValueError(f"La hoja {sheet.title} no tiene filtros o encabezado congelado")
        event_formulas = [cell.value for cell in workbook["Eventos"]["O"][1:]]
        if event_formulas and not all(str(value).startswith("=COUNTIF(") for value in event_formulas):
            raise ValueError("Las fórmulas de número de fuentes no son válidas")
        workbook.close()

    def _write_events(self, sheet: Worksheet, rows: Sequence[object], publications_end: int) -> None:
        headers = [
            "ID", "Código", "Título", "Resumen", "Tipo de evento", "Estado", "Municipio",
            "Inicio", "Última actualización", "Estado del evento", "Prioridad", "Puntaje",
            "Revisión", "Actores", "Número de fuentes", "Motivos del puntaje",
        ]
        values: list[list[object]] = []
        for index, row in enumerate(rows, start=2):
            values.append([
                row["id"], row["event_code"], row["title"], row["description"], row["event_type"],
                row["state"], row["municipality"], self._excel_datetime(row["start_date"]),
                self._excel_datetime(row["last_update"]), row["status"], row["importance_level"],
                row["priority_score"], self._yes_no(row["needs_review"]),
                str(row["actors"] or "").replace(",", ", "),
                f"=COUNTIF('Publicaciones'!$B$2:$B${publications_end},B{index})",
                self._json_list(row["score_reasons"]),
            ])
        self._populate(sheet, headers, values, "EventosTable")
        self._set_widths(sheet, [8, 25, 48, 65, 32, 20, 20, 19, 19, 18, 13, 10, 11, 32, 17, 55])
        for cell in list(sheet["H"])[1:] + list(sheet["I"])[1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
        last_row = sheet.max_row
        sheet.conditional_formatting.add(
            f"K2:K{last_row}", FormulaRule(formula=["K2=\"Alto\""], fill=HIGH_FILL)
        )
        sheet.conditional_formatting.add(
            f"K2:K{last_row}", FormulaRule(formula=["K2=\"Crítico\""], fill=HIGH_FILL)
        )
        sheet.conditional_formatting.add(
            f"K2:K{last_row}", FormulaRule(formula=["K2=\"Medio\""], fill=MEDIUM_FILL)
        )

    def _write_publications(self, sheet: Worksheet, rows: Sequence[object]) -> None:
        headers = [
            "ID", "Código del evento", "Fuente", "Título", "Fecha de publicación", "Estado",
            "Municipio", "Partido", "Tipo de evento", "Procesamiento", "Revisión", "Relación",
            "Es fuente primaria", "Duplicada", "URL", "Autor",
        ]
        values = [[
            row["id"], row["event_code"], row["source_name"], row["title"],
            self._excel_datetime(row["published_at"]), row["state_detected"],
            row["municipality_detected"], row["party_detected"], row["event_type_detected"],
            row["status"], self._yes_no(row["needs_review"]), row["relationship_type"],
            self._yes_no(row["is_primary_source"]), self._yes_no(row["is_duplicate"]),
            row["url"], row["author"],
        ] for row in rows]
        self._populate(sheet, headers, values, "PublicacionesTable")
        self._set_widths(sheet, [8, 25, 27, 55, 20, 20, 20, 23, 32, 18, 11, 22, 18, 12, 60, 25])
        for cell in list(sheet["E"])[1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
        self._format_links(sheet, "O")

    def _write_actors(self, sheet: Worksheet, rows: Sequence[object]) -> None:
        headers = ["Código del evento", "Evento", "Actor", "Tipo", "Partido", "Estado", "Prioritario"]
        values = [[
            row["event_code"], row["event_title"], row["name"], row["actor_type"],
            row["party"], row["state"], self._yes_no(row["is_priority"]),
        ] for row in rows]
        self._populate(sheet, headers, values, "ActoresTable")
        self._set_widths(sheet, [25, 55, 30, 18, 23, 20, 13])

    def _write_review(self, sheet: Worksheet, rows: Sequence[object]) -> None:
        headers = ["ID", "Título", "Fuente", "Estado", "Tipo de evento", "Motivos de revisión", "URL"]
        values = [[
            row["id"], row["title"], row["source_name"], row["state_detected"],
            row["event_type_detected"], self._json_list(row["review_reasons"]), row["url"],
        ] for row in rows]
        self._populate(sheet, headers, values, "RevisionTable")
        self._set_widths(sheet, [8, 55, 27, 20, 32, 60, 60])
        self._format_links(sheet, "G")
        if rows:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.fill = REVIEW_FILL

    def _write_sources(self, sheet: Worksheet, rows: Sequence[object], publications_end: int) -> None:
        headers = [
            "ID", "Fuente", "Tipo", "Estado base", "Confiabilidad", "Activa", "URL base",
            "Publicaciones recopiladas",
        ]
        values: list[list[object]] = []
        for index, row in enumerate(rows, start=2):
            values.append([
                row["id"], row["name"], row["source_type"], row["state"],
                row["reliability_level"], self._yes_no(row["is_active"]), row["base_url"],
                f"=COUNTIF('Publicaciones'!$C$2:$C${publications_end},B{index})",
            ])
        self._populate(sheet, headers, values, "FuentesTable")
        self._set_widths(sheet, [8, 30, 18, 20, 15, 10, 60, 25])
        self._format_links(sheet, "G")

    def _populate(
        self,
        sheet: Worksheet,
        headers: list[str],
        rows: Iterable[list[object]],
        table_name: str,
    ) -> None:
        sheet.append(headers)
        written = 0
        for row in rows:
            sheet.append(row)
            written += 1
        if not written:
            sheet.append(["Sin registros", *([None] * (len(headers) - 1))])
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 32
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{self._column_letter(len(headers))}{sheet.max_row}"
        sheet.sheet_view.showGridLines = False
        table = Table(displayName=table_name, ref=sheet.auto_filter.ref)
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE, showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)

    @staticmethod
    def _set_widths(sheet: Worksheet, widths: list[float]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[ExcelReportExporter._column_letter(index)].width = width

    @staticmethod
    def _format_links(sheet: Worksheet, column: str) -> None:
        for cell in list(sheet[column])[1:]:
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT

    @staticmethod
    def _excel_datetime(value: str | None) -> datetime | None:
        if not value:
            return None
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None)

    @staticmethod
    def _yes_no(value: object) -> str:
        return "Sí" if bool(value) else "No"

    @staticmethod
    def _json_list(value: str | None) -> str:
        if not value:
            return ""
        try:
            items = json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return str(value)
        return " · ".join(str(item) for item in items)

    @staticmethod
    def _column_letter(index: int) -> str:
        result = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
