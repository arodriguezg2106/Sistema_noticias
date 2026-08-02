from openpyxl import load_workbook

from app.collectors import MockCollector
from app.exports import ExcelReportExporter
from app.pipeline import ProcessingPipeline
from app.repositories import Database


def test_excel_export_contains_auditable_sheets(tmp_path, project_root, configs) -> None:
    database = Database(tmp_path / "excel.db")
    pipeline = ProcessingPipeline(database, configs)
    pipeline.initialize()
    pipeline.run(MockCollector(project_root / "data" / "mock_publications.json"))

    output = ExcelReportExporter(database).export(tmp_path / "reporte.xlsx")
    workbook = load_workbook(output, data_only=False)

    assert workbook.sheetnames == [
        "Eventos", "Publicaciones", "Actores", "Revisión manual", "Fuentes"
    ]
    assert workbook["Eventos"].max_row == 5
    assert workbook["Publicaciones"].max_row == 7
    assert workbook["Eventos"]["O2"].value.startswith("=COUNTIF(")
    assert workbook["Publicaciones"]["O2"].hyperlink is not None
    assert all(sheet.freeze_panes == "A2" for sheet in workbook.worksheets)
    assert all(sheet.auto_filter.ref for sheet in workbook.worksheets)
    workbook.close()
